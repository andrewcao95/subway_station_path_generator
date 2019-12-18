from itertools import chain
from tqdm import tqdm
import time
import openpyxl

minChangeRes = []
shortestPathRes = []
bestPathRes = []


def buildSubway(lines):
    for key in lines.keys():
        value = lines[key]
        lines[key] = value.split()

    stations = set()
    for key in lines.keys():
        stations.update(set(lines[key]))

    system = {}
    for station in stations:
        next_station = {}
        for key in lines:
            if station in lines[key]:
                line = lines[key]
                idx = line.index(station)
                if idx == 0:
                    next_station[line[1]] = key
                elif idx == len(line) - 1:
                    next_station[line[idx - 1]] = key
                else:
                    next_station[line[idx - 1]] = key
                    next_station[line[idx + 1]] = key
        system[station] = next_station
    return system


def updateSubway(originSubway):
    originSubway[u'西直门'][u'积水潭'] = 'line_2'
    originSubway[u'积水潭'][u'西直门'] = 'line_2'
    originSubway[u'劲松'][u'潘家园'] = 'line_10'
    originSubway[u'潘家园'][u'劲松'] = 'line_10'
    return originSubway


def shortestPathPlaner(start, goal):
    if start == goal:
        return [start]

    explored = set()
    queue = [[start]]

    while queue:
        path = queue.pop(0)
        s = path[-1]
        for state, action in revisedSubway[s].items():
            # print(revisedSubway[s].items())
            if state not in explored:
                explored.add(state)
                path2 = path + [action, state]
                if state == goal:
                    return path2
                else:
                    queue.append(path2)
    return []


def minChangePathPlaner(start, goal):
    if start == goal:
        return [start]
    explored = set()
    queue = [[start, ('', 0)]]
    while queue:
        path = queue.pop(0)

        s = path[-2]
        linenum, changetimes = path[-1]

        if s == goal:
            return path[: -1]
        for state, action in revisedSubway[s].items():
            if state not in explored:
                linechange = changetimes
                explored.add(state)
                if linenum != action:
                    linechange += 1
                path2 = path[:-1] + [action, state, (action, linechange)]
                queue.append(path2)
                queue.sort(key=lambda path: path[-1][-1])
    return []


def get_result(type, start, end):
    if type == "best":
        info = bestPathPlannerRes(start, end)
    elif type == "shortest":
        info = shortestPathPlanerRes(start, end)
    elif type == "minChange":
        info = minChangePathPlanerRes(start, end)

    return "\n出发站:" + start + "\n目的站：" + end + "\n间隔站点数：" + str(info[0]) + "\n换乘次数：" + str(info[1]) + "\n规划路径：" + str(
        info[2])


def generate_summary_result(type):
    stations = []
    for item in subway_lines.keys():
        stations.append(subway_lines[item])
    stations = list(chain.from_iterable(stations))

    stations = stations[:50] # for test

    book = openpyxl.Workbook()
    if type == "best":
        string1 = "station_num_best"
        string2 = "change_num_best"
        planner = bestPathPlannerRes
    elif type == "shortest":
        string1 = "station_num_shortest"
        string2 = "change_num_shortest"
        planner = shortestPathPlanerRes
    elif type == "minChange":
        string1 = "station_num_min_change"
        string2 = "change_num_min_change"
        planner = minChangePathPlanerRes

    sheet1 = book.create_sheet(string1)
    sheet2 = book.create_sheet(string2)
    sheet = book['Sheet']
    book.remove(sheet)

    print('\n-------------正在生成 [ ' + string1 + '] 站点数关系表----------\n')
    for i in range(len(stations)):
        sheet1.cell(i + 2, 1).value = stations[i]
        sheet1.cell(1, i + 2).value = stations[i]

    for x in tqdm(range(len(stations))):
        for y in range(len(stations)):
            station_num, change_num, res = planner(stations[x], stations[y])
            sheet1.cell(x + 2, y + 2, station_num)

    print('\n-------------正在生成 [ ' + string2 + '] 站点数关系表----------\n')
    for i in range(len(stations)):
        sheet2.cell(i + 2, 1).value = stations[i]
        sheet2.cell(1, i + 2).value = stations[i]

    for x in tqdm(range(len(stations))):
        for y in range(len(stations)):
            station_num, change_num, res = planner(stations[x], stations[y])
            sheet2.cell(x + 2, y + 2, change_num)

    file = './' + type + '_result.xlsx'
    book.save(file)
    print(file + '文件生成完成\n')


def minChangePathPlanerRes(here, there):
    res = minChangePathPlaner(here, there)
    path = res[::2]
    temp = res[1::2]
    count = 0
    for i in range(len(temp) - 1):
        if temp[i] != temp[i + 1]:
            count += 1
    return len(path) - 1, count, res


def shortestPathPlanerRes(here, there):
    res = shortestPathPlaner(here, there)
    path = res[::2]
    temp = res[1::2]
    count = 0
    for i in range(len(temp) - 1):
        if temp[i] != temp[i + 1]:
            count += 1
    return len(path) - 1, count, res


def bestPathPlannerRes(here, there):
    minChangeRes = minChangePathPlaner(here, there)
    shortestRes = shortestPathPlaner(here, there)
    if (len(shortestRes) >= len(minChangeRes)):
        res = minChangeRes
    else:
        res = shortestRes
    path = res[::2]
    temp = res[1::2]
    count = 0
    for i in range(len(temp) - 1):
        if temp[i] != temp[i + 1]:
            count += 1
    return len(path) - 1, count, res


if __name__ == '__main__':
    subway_lines = {}

    # 加载站点数据
    with open("./station_data_new.txt", "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines:
            line_name = line.split(":")[0].strip()
            line_stations = line.split(":")[1].strip()
            subway_lines[line_name] = line_stations

    originSubway = buildSubway(subway_lines)
    revisedSubway = updateSubway(originSubway)

    # 提示输入
    input_type = input("\n请选择生成模式 [1]站点间路径，[2]excel汇总表格生成：")

    if input_type == "1":
        input_station_start_info = input("请输入出发站名称（例如：北京大学东门）：")
        input_station_end_info = input("请输入目的站名称（例如：传媒大学）：")
        input_type_info = input('请输入待生成的路径类型 [1]最佳路径（在[2]和[3]之间择优选择）, [2]站点间隔最短， [3]换乘线路最少（例如：输入1，2，3数字）：')

        res = ""
        if input_type_info == "1":
            res = get_result("best", input_station_start_info, input_station_end_info)
            print("您选择的是[1]最佳路径")
        elif input_type_info == "2":
            res = get_result("shortest", input_station_start_info, input_station_end_info)
            print("您选择的是[2]站点间隔最短")
        elif input_type_info == "3":
            res = get_result("minChange", input_station_start_info, input_station_end_info)
            print("您选择的是[3]换乘线路最少")
        else:
            print('您的输入有误，请重新运行程序并输入（提示：输入1，2，3数字）')
        print(res)

    elif input_type == "2":
        print("正在预加载检索数据，请等待.......\n")
        generate_summary_result("best")
        time.sleep(2)
        generate_summary_result("shortest")
        time.sleep(2)
        generate_summary_result("minChange")
